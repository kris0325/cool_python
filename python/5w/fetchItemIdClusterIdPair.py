# -*- coding:utf-8 -*-
import  xdrlib ,sys
import xlrd
import xlwt
from openpyxl import Workbook

#####
# 1. get productid itemid from redshift                           --> result1 = PidAndIid.xlsx
# 2. get pid_iid from result1. CreatePidAndIid.py                 --> result2 = Pid_Iid.xlsx
# 3. get pid_iid clusterid from result2 by quert hive             --> result3 = query_result.xlsx
# 4. get itemid clusterid from result3  GetItemIdAndClusterId.py  --> result3 = itemIdAndClusterId.xlsx
#####


wb = Workbook()
ws = wb.active

def open_excel(file):
    try:
        data = xlrd.open_workbook(file)
        return data
    except Exception,e:
        print str(e)

def getItemIdAndClusterId(by_name):
    #data1 = open_excel('itemIdAndClusterId.xlsx')
    #data2 = open_excel('itemId5w.xlsx')
    data1 = open_excel('itemIdAndClusterId.xlsx')
    data2 = open_excel('itemId5w.xlsx')

    table1 = data1.sheet_by_name(by_name)
    nrows1 = table1.nrows
    table2 = data2.sheet_by_name(by_name)
    nrows2 = table2.nrows

    r =1
    for i in range(1,nrows2):
        for j in range(1,nrows1):
            if table2.cell_value(i,0) == table1.cell_value(j,0):
                ws.cell(row=r,column=1, value=table2.cell_value(i,0))
                ws.cell(row=r,column=2, value=table1.cell_value(j,1))
                r = r+1
                break


        # get itemId from pid_iid
        #ws.cell(row=i,column=1,value= str[str.find(s)+1:-1])
        #get clusterid
        #ws.cell(row=i,column=2,value=table.cell_value(i,1))
    #pass



def main():
      getItemIdAndClusterId(by_name = u'Sheet1')
      wb.save('ICPair.xlsx')

if __name__=="__main__":
    main()

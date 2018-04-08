# -*- coding:utf-8 -*-
import  xdrlib ,sys
import xlrd
import xlwt
from openpyxl import Workbook

#####
# 1. get productid itemid from redshift                           --> result1 = PidAndIid.xlsx
# 2. get pid_iid from result1. CreatePidAndIid.py                 --> result2 = Pid_Iid.xlsx
# 3. get pid_iid clusterid from result2 by quert hive             --> result3 = query_result.xlsx
# 4. get itemid clusterid from result3  GetItemIdAndClusterId.py  --> result3 = itemIdAndClusterId.xlsx
#####


wb = Workbook()
ws = wb.active

def open_excel(file):
    try:
        data = xlrd.open_workbook(file)
        return data
    except Exception,e:
        print str(e)

def getItemIdAndClusterId(file, by_name):
    data = open_excel(file)
    table = data.sheet_by_name(by_name)
    nrows = table.nrows
    s = 'I'
    for i in range(1,nrows):
        str = '\'' + table.cell_value(i,0) + '\''

        # get itemId from pid_iid
        ws.cell(row=i,column=1,value= str[str.find(s)+1:-1])
        ws.cell(row=i,column=2,value=table.cell_value(i,1))
    #pass



def main():
      getItemIdAndClusterId(file = 'query_result.xlsx', by_name = u'Sheet')
      wb.save('itemIdAndClusterId.xlsx')

if __name__=="__main__":
    #getItemIdAndClusterId to save to redis
    main()

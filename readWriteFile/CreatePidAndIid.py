# -*- coding: utf-8 -*-
import  xdrlib ,sys
import xlrd
import xlwt
from openpyxl import Workbook

wb = Workbook()
#选择第一个工作表
ws = wb.active
#创建一个新的工作表
#ws1 = wb.create_sheet('Mysheet')

workbook = xlwt.Workbook(encoding = 'ascii')
worksheet = workbook.add_sheet('Sheet1',cell_overwrite_ok=True)

def open_excel(file= 'PidAndIid.xlsx'):
    try:
        data = xlrd.open_workbook(file)
        return data
    except Exception,e:
        print str(e)

#join col by "-" 构造pid_Iid lists
def join_clo(file= 'PidAndIid.xlsx',colnameindex=0,by_name=u'Sheet1'):
    data = open_excel(file)
    table = data.sheet_by_name(by_name)
    nrows = table.nrows #行数
    ncols = table.ncols #列数
    colnames = table.row_values(colnameindex) #某一行数据
    joinstr = '_'
    lists = []
    for i in range(0,nrows):
             newcell = '\''+'P'+ str(int(table.cell_value(i,0))) + '_'+ 'I'+ str(int(table.cell_value(i,1))) +'\''+','
             # xlwt 只支持写入.xls格式文件，即07版之前的Excel，并且生成Excel的单个sheet最大行数限制为65535行
             #worksheet.write(i, 0, label = newcell)
             ws.cell(row=i+1,column=1,value=newcell)

def main():
  join_clo()
  #workbook.save('Excel_Workbook.xls')
  wb.save('Pid_Iid.xlsx')

if __name__=="__main__":
    main()
